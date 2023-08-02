# Importing required module
import openpyxl

def process_spreadsheet(file_path, search_key):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    processed_data = []
    colNum = 1
    if any(char.isdigit() for char in search_key):
        colNum = 2
    # search for the search key in the spreadsheet
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, colNum)
        # if cell.value == search_key:
        if search_key.lower() in str(cell.value).lower():
            # get the values of the cells in the row
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                processed_data.append(cell.value)

    return processed_data

if __name__ == "__main__":
    spreadsheet_path = "sample.xlsx"
    search_key = input("Enter the name or id to search: ")

    results = process_spreadsheet(spreadsheet_path, search_key)

    if results:
        print("Processed Data:")
        for result in results:
            print(result)
    else:
        print("No results found")

