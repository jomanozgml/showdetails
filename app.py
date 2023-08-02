# Importing required module
from flask import Flask, render_template, request, jsonify
import openpyxl

app = Flask(__name__)

def process_spreadsheet(file_path, search_key):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    header = []
    for cell in sheet[1]:
        header.append(cell.value)
    processed_data = []
    colNum = 1
    if any(char.isdigit() for char in search_key):
        colNum = 2
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, colNum)
        if search_key.lower() in str(cell.value).lower():
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                processed_data.append(cell.value)
            break

    return [processed_data, header]

@app.route('/')

def index():
    return render_template('index.html')

@app.route('/search', methods=['POST'])

def search():
    search_key = request.form['search_key']
    spreadsheet_path = "sample.xlsx"
    results_data = process_spreadsheet(spreadsheet_path, search_key)
    return render_template('index.html', search_key=search_key, results=results_data[0], header=results_data[1])

if __name__ == "__main__":
    app.run(debug=True)

